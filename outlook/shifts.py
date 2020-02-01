#!/usr/bin/env python3

import os
import sys
import time
import datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

shift_path = "\\\iuser\\hubufsr\\ROC\\GCSO\\Special\\Admin\\Shifts\\"
shift = "AMS-SDD schedule 2020.xlsx"

def open_schedule(f_path, file):
    
    if os.path.isfile(f_path + file):
        print("\n'%s' is valid file" % (file))
        print("Reading excel...\n")
    elif os.path.isfile(f_path + file) == False:
        print("\n'%s' does not exist or path is invalid\n\n" % (file))


def get_shifts():

    # clear screen on execution
    os.system('clear')
    # validate the shift file path
    open_schedule(shift_path, shift)

    # create excel object with path + file 
    wb = openpyxl.load_workbook(shift_path + shift)
    # select the correct sheet
    sheet = wb["AMS_2020"]
    print(f"sheet: {sheet.title}")

    # get current date
    current_date = datetime.datetime.now().strftime("%d/%m/%Y")
    print(f"current date: {current_date}")

    cell_column = ""

    # for every row of cell object between P2 and BD2, iterate over the cell objects in the row
    for r in sheet["P2":"OJ2"]:
        for c in r:
            # format the datetime.datetime object into "Day/Month/Year"
            date = c.value.strftime("%d/%m/%Y")
            # if current date == date in excel, save the value to variable
            if current_date == date:
                #print(c.row, c.column, c.coordinate)
                print(f"found date in excel: {date}")
                print("saving today's shifts...")

                # save the column to cell_column
                cell_column = c.column

    shifts_on_current_date = [sheet.cell(row=i, column=cell_column).value for i in range(4, 22)]
    agents = [sheet.cell(row=i, column=2).value for i in range(4, 22)]
    #print(shifts_on_current_date)
    #print(agents)

    return shifts_on_current_date, agents

current_shifts, agent_list = get_shifts()

def create_dictionary(c_shifts, a_list):

    # create a dict with key value pairs => "Agent" : "shift"
    agents_shift_today = dict(zip(a_list, c_shifts))
    #print(agents_shift_today)
    return agents_shift_today