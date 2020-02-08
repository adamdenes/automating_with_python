#!/usr/bin/env python3

import os
import sys
import time
import logging
import ansicon
import datetime
import openpyxl
import win32com.client
from openpyxl.utils import get_column_letter, column_index_from_string

logfile = "C:\\Users\\612561487\\desktop\\projects\\outlook\\" + str(datetime.datetime.now().strftime('%Y_%m_%d.txt'))
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(name)-5s %(levelname)-5s %(message)s",
                    datefmt="[%Y.%m.%d %H:%M:%S]",
                    filename=logfile,
                    filemode="a")

start_time = start_time = time.time()
ansicon.load()
shift_path = "\\\iuser\\hubufsr\\ROC\\GCSO\\Special\\Admin\\Shifts\\"
shift = "AMS-SDD schedule 2020.xlsx"


def open_schedule(f_path, file):
    date_time = datetime.datetime.now().strftime("%Y.%m.%d %H:%M:%S")
    logging.info(f"************* cycle started *************")
    #print(f"Operation started at {date_time}")

    if os.path.isfile(f_path + file):
        #print("\n\u001b[33m'%s'\u001b[0m is valid file" % (file))
        #print("Reading excel...\n")
        logging.info("'%s' is valid file" % (file))
        logging.info("Reading excel...")
    elif os.path.isfile(f_path + file) == False:
        #print("\n\u001b[33m'%s'\u001b[0m does not exist or path is invalid\n\n" % (file))
        logging.info("'%s' does not exist or path is invalid" % (file))

       
def get_shifts():

    # clear screen on execution
    os.system("cls")
    # validate the shift file path
    open_schedule(shift_path, shift)

    # create excel object with path + file 
    wb = openpyxl.load_workbook(shift_path + shift)
    # select the correct sheet
    sheet = wb["AMS_2020"]
    #print(f"sheet: \u001b[33m{sheet.title}\u001b[0m")

    # get current date
    current_date = datetime.datetime.now().strftime("%d/%m/%Y")
    #print(f"current date: \u001b[33m{current_date}\u001b[0m")

    cell_column = ""

    # for every row of cell object between P2 and BD2, iterate over the cell objects in the row
    for r in sheet["P2":"OJ2"]:
        for c in r:
            # format the datetime.datetime object into "Day/Month/Year"
            date = c.value.strftime("%d/%m/%Y")
            # if current date == date in excel, save the value to variable
            if current_date == date:
                #print(c.row, c.column, c.coordinate)
                #print(f"found date in excel: \u001b[33m{date}\u001b[0m")
                #print("saving today's shifts...\n")
                logging.info("saving today's shifts...")
                # save the column to cell_column
                cell_column = c.column

    shifts_on_current_date = [sheet.cell(row=i, column=cell_column).value for i in range(4, 22)]
    agents = [sheet.cell(row=i, column=2).value for i in range(4, 22)]

    return shifts_on_current_date, agents

current_shifts, agent_list = get_shifts()


def create_dictionary(c_shifts, a_list):

    now = datetime.datetime.now().hour

    # create a dict with key value pairs => "Agent" : "shift"
    agents_shift_today = dict(zip(a_list, c_shifts))
    present_dict = {}
    absent_dict = {}

    #print("working today:\n")

    for key, value in agents_shift_today.items():
        # if the shift is AL, BH, *, or empty -> save it to 'absent'
        if value == "AL" or value == "BH" or value == "*" or value == None:
            absent_dict[key] = value        
        #elif (value == "M8" and now > 14) or (value == "D8" and now > 17) or (value == "D12" and (6 <= now or now < 18)) or (value == "A8" and (14 <= now or now < 22)) or (value == "N8" and (now >= 22 or now < 6)) or (value == "N12" and (18 <= now or now < 6)):
        #    absent_dict[key] = value
        else:
            # else save the present agents into 'present'
            present_dict[key] = value
            #print(f"\t\u001b[33m*\u001b[0m {key} \u001b[32m->\u001b[0m \u001b[36m{value}\u001b[0m")

    #print("\n", absent_dict, "\n")
    #print(present_dict, "\n")
    return absent_dict, present_dict

absent, present = create_dictionary(current_shifts, agent_list)

# create outlook object
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
folders = outlook.Folders

inboxes = [folders[folder].Name for folder in range(folders.Count)]
boxes = [i for i in inboxes if "adam" not in i]
#print(f"\n{boxes}\n")


def generate_mailboxes(box_lst):

    for box in box_lst:
        root_f = outlook.Folders[box]
        sub_f = root_f.Folders["Inbox"]

        if str(root_f.Name) == "Servicedesk Debrecen G":
            personal_inboxes = sub_f.Folders["Team's folders"]
            processed = sub_f.Folders["Processed"]
            #planned_work = sub_f.Folders["Planned works"]
        elif str(root_f.Name) == "CSC-NL-PHILIPS G":
            personal_inboxes = sub_f.Folders["+++TEAM+++"]
            processed = sub_f.Folders["+++PROCESSED+++"]
            #planned_work = sub_f.Folders["+++PLANNED WORKS+++"]
        else:
            personal_inboxes = sub_f.Folders["+++TEAM+++"]
            processed = sub_f.Folders["+++PROCESSED+++"]
            planned_work = processed.Folders["+++Planned Works+++"]
        personal_inb = personal_inboxes.Folders

        msgs = sub_f.Items
        msg = msgs.GetFirst()
    # took out planned_work
    return root_f, sub_f, personal_inboxes, processed, personal_inb, msgs, msg


CSC = generate_mailboxes([boxes[0]])
SDD = generate_mailboxes([boxes[1]])
AMS = generate_mailboxes([boxes[2]])
logging.info(f"Checking mailboxes: {boxes}")

def move_mails(item_in_mailb):

    #pw = item_in_mailb[-3]
    sender = item_in_mailb[0]
    processed = item_in_mailb[3]

    for message in list(item_in_mailb[-2]):
        categories = message.categories
        categ = categories.split(',')

        mail_arrived = datetime.datetime.strftime(message.CreationTime, "%Y.%m.%d %H:%M:%S")
        
        # CISCO MERAKI MAILS
        #if str(message.Sender) == "Cisco Meraki - No Reply":
        #    print(f"moving \u001b[33m{message.Sender}\u001b[0m to folder => \u001b[32m{processed}\u001b[0m")
        #    message.Move(processed)
        #    time.sleep(1)
        #if "Change Creation" in str(message.Subject) or "Change Closure" in str(message.Subject):
        #    print(str(message.Subject))
        #    message.Move(pw)
        #elif "Scheduled Maintenance" in str(message.Subject) or "Planned Work" in str(message.Subject):
        #    print(str(message.Subject))
        #    message.Move(pw)

        logging.info("waiting for email flagging...")
        if not categories:
            continue
        elif len(categ) > 4:
            continue
        elif str(message.Sender) == str(sender):
            #print(sender, message.Sender)
            #print(f"ticket: \u001b[33m{categ[0].strip()}\u001b[0m => \u001b[32m{processed}\u001b[0m")
            #print(f"\tmoving \u001b[33m{categ[0].strip()}\u001b[0m to folder => \u001b[35m{sender}\u001b[0m - \u001b[32m{processed}\u001b[0m")            
            logging.info(f"moving {categ[0].strip()} to folder => {sender} - {processed}")
            message.Move(processed)
        else:
            #print(f"ticket: \u001b[33m{categ[0].strip()}\u001b[0m => assignee: \u001b[32m{categ[1].strip()}\u001b[0m => mail arrived at: \u001b[36m{mail_arrived}\u001b[0m")
            logging.info(f"{categ[0].strip()} - {categ[1].strip()}")
            for p_sub in item_in_mailb[4]:
                if (str(categ[1].strip()) in str(p_sub)) and (not str(categ[1].strip()) in absent):
                    #print(f"\tmoving \u001b[33m{categ[0].strip()}\u001b[0m to folder => \u001b[35m{sender}\u001b[0m - \u001b[32m{p_sub}\u001b[0m")
                    logging.info(f"moving {categ[0].strip()} to folder => {sender} - {p_sub}")
                    message.Move(p_sub)
                    messages = item_in_mailb[1].Items
                    message = messages.GetFirst()
                    time.sleep(1)
        
def main(mailboxes):

    for mailb in mailboxes:
        move_mails(mailb)


if __name__ == "__main__":
    main([SDD, AMS, CSC])
    logging.info(f"************* cycle finished in {time.time() - start_time} *************")